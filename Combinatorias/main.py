from math import factorial
from itertools import combinations, permutations
from openpyxl import Workbook, load_workbook
from os.path import exists
from datetime import datetime


def to_save(date, operation, set_, operation_result, new_sets, filename='registro_operaciones.xlsx'):
    if exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Operaciones'
        ws.append([
            'Fecha y hora',
            'Operación',
            'Conjunto original',
            'Resultado',
            'Nuevos conjuntos',
        ])

    detalle = ', '.join([''.join(x) for x in new_sets])

    ws.append([
        date,
        operation,
        ', '.join(set_),
        operation_result,
        detalle
    ])

    wb.save(filename)


def _combination(set_, num_combinations):
    operation = factorial(len(set_)) // factorial(num_combinations) * (
                factorial(len(set_)) - factorial(num_combinations))
    total_combinations = list(combinations(set_, num_combinations))

    return [operation, total_combinations]


def _permutation(set_):
    operation = factorial(len(set_))

    total_permutations = list(permutations(set_))
    return [operation, total_permutations]


def _variation(set_, num_variations):
    operation = factorial(len(set_)) // (factorial(len(set_)) - factorial(num_variations))
    total_variations = list(permutations(set_, num_variations))
    return [operation, total_variations]


"""
    se ingresa un conjunto sin operadores a los costados 

        Ej.
            a,b,c
    original_set = original_set.split(',') <---- Lo separa y guarda en una lista

    al final guarda en un excel el resultado y las permutaciones/combinaciones/variaciones
    que se pueden realizar
"""

original_set = input('Ingresa el conjunto de numeros: ')
original_set = original_set.split(',')

status = input(
    'Selecciona opcion:\n'
    'a - combinacion\n'
    'b - permutacion\n'
    'c - variaciones\n'
)

operation_name = 'combinacion' if status == 'a' else 'permutacion' if status == 'b' else 'variaciones'
result = 0
list_set = []
if status == 'a' or status == 'c':
    selection_numbers = int(input('Numero de elementos elegidos:'))
    result, list_set = _combination(original_set, selection_numbers) if status == 'a' else _variation(original_set,
                                                                                                      selection_numbers)

elif status == 'b':
    result, list_set = _permutation(original_set)
else:
    print('Opcion invalida')
print('Resultado:', result)
fecha_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
to_save(fecha_hora, operation_name, original_set, result, list_set)





